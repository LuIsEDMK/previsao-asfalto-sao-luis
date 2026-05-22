"""
Projeto: Preditor de Falhas em Vias - Sao Luis / MA
Script:  import_os_semusc.py
Tarefa:  Ler template_os_semusc.xlsx preenchido pela SEMUSC
         e converter para o formato padrão do projeto
         Gera: 02_repair_history_real.csv (NÃO sobrescreve o simulado)
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

import os
import warnings
from datetime import date, datetime

import pandas as pd
from openpyxl import load_workbook

COD_DIR = r"w:\projetos vscode\projetos para prefeitura\1projeto\codigos"
CSV_DIR = r"w:\projetos vscode\projetos para prefeitura\csv"

TEMPLATE = f"{COD_DIR}/template_os_semusc.xlsx"
SAIDA    = f"{CSV_DIR}/02_repair_history_real.csv"
HOJE     = date.today()

def sep(t=""):
    if t:
        print(f"\n{'='*60}\n  {t}\n{'='*60}")

# ─── Mapeamentos português → código do projeto ────────────────────
MAP_TIPO_SERVICO = {
    "Tapa-buraco":                  "tapa-buraco",
    "Recapeamento parcial":         "recapeamento parcial",
    "Recapeamento total":           "recapeamento total",
    "Reposição de paralelepípedo":  "reposição paralelepípedo",
    "Correção de erosão":           "outros",
    "Outro":                        "outros",
}

MAP_TIPO_PROBLEMA = {
    "Buraco":                    "buraco",
    "Trinca / rachadura":        "trinca",
    "Afundamento":               "afundamento",
    "Desgaste da superfície":    "desgaste",
    "Deslocamento de pedras":    "deslocamento",
    "Erosão / escorregamento":   "erosao",
    "Outro":                     "outros",
}

MAP_COMO_IDENTIFICADO = {
    "Reclamação da ouvidoria":    "ouvidoria",
    "Reclamação pelo WhatsApp":   "whatsapp",
    "Reclamação pelo Colab App":  "colab_app",
    "Vistoria da equipe":         "vistoria",
    "Emergência / acidente":      "emergencia",
    "Outro":                      "outros",
}

MAP_PRIORIDADE = {
    "Urgente (risco de acidente)":    "urgente",
    "Alta (via principal)":           "alta",
    "Média (via secundária)":         "media",
    "Baixa (manutenção preventiva)":  "baixa",
}


# ─── Utilitários ──────────────────────────────────────────────────

def converter_data(val):
    """
    Converte valor de célula Excel para date.
    Aceita: date/datetime, string DD/MM/AAAA ou AAAA-MM-DD, None.
    """
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    if not s or s.lower() in ("none", "nan", ""):
        return None
    # Tenta DD/MM/AAAA
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def converter_numero(val, tipo=float):
    """Converte valor de célula para número, retorna None se inválido."""
    if val is None:
        return None
    try:
        return tipo(str(val).replace(",", ".").replace("R$", "").strip())
    except (ValueError, AttributeError):
        return None


def mapear(val, mapa, campo):
    """Aplica mapeamento; retorna valor original em minúsculo se não encontrado."""
    if not val:
        return None
    val_str = str(val).strip()
    resultado = mapa.get(val_str)
    if resultado is None:
        # Tentativa case-insensitive
        for k, v in mapa.items():
            if k.lower() == val_str.lower():
                return v
        return val_str.lower()
    return resultado


# ─── Verificar arquivo ────────────────────────────────────────────

sep("Verificando arquivo de entrada")

if not os.path.exists(TEMPLATE):
    print(f"  [ERRO] Arquivo nao encontrado: {TEMPLATE}")
    print(f"  Execute primeiro: py -X utf8 create_template_semusc.py")
    sys.exit(1)

print(f"  [OK] Arquivo encontrado: {TEMPLATE}")

# ─── Ler o template ───────────────────────────────────────────────

sep("Lendo template Excel")

try:
    wb = load_workbook(TEMPLATE, data_only=True)
except Exception as e:
    print(f"  [ERRO] Nao foi possivel abrir o arquivo: {e}")
    sys.exit(1)

ABA_NOME = "📋 PREENCHER AQUI"
if ABA_NOME not in wb.sheetnames:
    print(f"  [ERRO] Aba '{ABA_NOME}' nao encontrada.")
    print(f"  Abas disponíveis: {wb.sheetnames}")
    sys.exit(1)

ws = wb[ABA_NOME]
print(f"  [OK] Aba '{ABA_NOME}' carregada")
print(f"  Dimensoes: {ws.dimensions}")

# ─── Processar linhas ─────────────────────────────────────────────

sep("Processando linhas")

# Linhas 1-4 = cabeçalho; dados a partir da linha 5
# (inclui exemplos 5-7 e dados reais 8+)

registros   = []
avisos      = []
rejeitadas  = []
n_lidas     = 0
n_vazias    = 0

for linha_num in range(5, ws.max_row + 1):
    # Ler células A até M (colunas 1-13); ignorar N (ID automático)
    celulas = [ws.cell(row=linha_num, column=col).value for col in range(1, 14)]

    # Pular linha completamente vazia (critério: col A vazia)
    if celulas[0] is None or str(celulas[0]).strip() == "":
        n_vazias += 1
        continue

    n_lidas += 1

    repair_id    = str(celulas[0]).strip()
    street_name  = str(celulas[1]).strip() if celulas[1] else ""
    bairro       = str(celulas[2]).strip() if celulas[2] else ""
    order_date   = converter_data(celulas[3])
    comp_date    = converter_data(celulas[4])
    tipo_srv     = mapear(celulas[5], MAP_TIPO_SERVICO,   "repair_type")
    tipo_prob    = mapear(celulas[6], MAP_TIPO_PROBLEMA,  "failure_type")
    custo        = converter_numero(celulas[7], float)
    crew         = converter_numero(celulas[8], int)
    como_id      = mapear(celulas[9], MAP_COMO_IDENTIFICADO, "reported_by")
    prio         = mapear(celulas[10], MAP_PRIORIDADE, "priority")
    material     = converter_numero(celulas[11], int)
    obs          = str(celulas[12]).strip() if celulas[12] else ""

    # Determinar status
    status = "concluido" if comp_date else "em_andamento"

    # ── Validações ───────────────────────────────────────
    erros_linha = []
    avs_linha   = []

    if not order_date:
        erros_linha.append("data_abertura inválida ou ausente")
    elif order_date > HOJE:
        avs_linha.append(f"data_abertura no futuro ({order_date})")

    if comp_date and order_date and comp_date < order_date:
        erros_linha.append(
            f"data_conclusao ({comp_date}) < data_abertura ({order_date})"
        )

    if custo is not None and custo < 0:
        erros_linha.append(f"custo negativo ({custo})")

    if not street_name:
        avs_linha.append("rua/avenida vazia")

    if not bairro:
        avs_linha.append("bairro vazio")

    if erros_linha:
        rejeitadas.append({
            "linha_excel": linha_num,
            "repair_id": repair_id,
            "motivo": " | ".join(erros_linha),
        })
        print(f"  [REJEITADA] Linha {linha_num} | {repair_id} — {' | '.join(erros_linha)}")
        continue

    if avs_linha:
        avisos.append({
            "linha_excel": linha_num,
            "repair_id": repair_id,
            "aviso": " | ".join(avs_linha),
        })
        print(f"  [AVISO]    Linha {linha_num} | {repair_id} — {' | '.join(avs_linha)}")

    registros.append({
        "repair_id":        repair_id,
        "segment_id":       "A_VINCULAR",
        "order_date":       order_date.strftime("%Y-%m-%d") if order_date else "",
        "completion_date":  comp_date.strftime("%Y-%m-%d") if comp_date else "",
        "repair_type":      tipo_srv or "",
        "failure_type":     tipo_prob or "",
        "repair_cost_brl":  round(custo, 2) if custo is not None else "",
        "crew_size":        int(crew) if crew is not None else "",
        "material_used_kg": int(material) if material is not None else "",
        "reported_by":      como_id or "",
        "priority":         prio or "",
        "status":           status,
        "bairro":           bairro,
        "street_name":      street_name,
        "notes":            obs,
    })

# ─── Relatório de importação ──────────────────────────────────────

sep("Relatorio de importacao")

n_ok  = len(registros)
n_avs = len(avisos)
n_rej = len(rejeitadas)

print(f"  Total de linhas lidas (nao vazias) : {n_lidas}")
print(f"  Importadas com sucesso             : {n_ok}")
print(f"  Com avisos (verificar)             : {n_avs}")
print(f"  Rejeitadas (erro grave)            : {n_rej}")
print(f"  Linhas vazias ignoradas            : {n_vazias}")

if avisos:
    print("\n  AVISOS (registros importados, mas verificar):")
    for av in avisos:
        print(f"    Linha {av['linha_excel']} | {av['repair_id']}: {av['aviso']}")

if rejeitadas:
    print("\n  REJEITADAS (nao importadas):")
    for rj in rejeitadas:
        print(f"    Linha {rj['linha_excel']} | {rj['repair_id']}: {rj['motivo']}")

# ─── Salvar CSV ───────────────────────────────────────────────────

if not registros:
    print("\n  [AV] Nenhum registro importado. Arquivo nao gerado.")
    print("  Verifique se o template foi preenchido na aba '📋 PREENCHER AQUI'.")
    sys.exit(0)

sep("Salvando 02_repair_history_real.csv")

df = pd.DataFrame(registros)

# Reordenar colunas para compatibilidade com 02_repair_history.csv
COLUNAS = [
    "repair_id", "segment_id", "order_date", "completion_date",
    "repair_type", "failure_type", "repair_cost_brl", "crew_size",
    "material_used_kg", "reported_by", "priority", "status",
    "bairro", "street_name", "notes",
]
for col in COLUNAS:
    if col not in df.columns:
        df[col] = ""
df = df[COLUNAS]

df.to_csv(SAIDA, index=False, encoding="utf-8")
print(f"  [OK] Arquivo gerado: {SAIDA}")
print(f"       {len(df)} registros salvos")

# Prévia dos primeiros registros
print("\n  Previa dos primeiros registros:")
print(df[["repair_id", "street_name", "bairro",
          "order_date", "repair_type", "priority"]].head(5).to_string(index=False))

# ─── Instrução final ──────────────────────────────────────────────
print(f"""
  [OK] Arquivo gerado: {SAIDA}

  Verifique os dados antes de substituir o arquivo simulado.
  Quando confirmar, renomeie manualmente para 02_repair_history.csv

  VERIFICACOES SUGERIDAS:
  1. Abrir 02_repair_history_real.csv no Excel ou editor de texto
  2. Confirmar que as datas estao no formato AAAA-MM-DD
  3. Confirmar que repair_type e priority estao em minúsculo
  4. Vincular segment_id: substituir 'A_VINCULAR' pelo ID do segmento
     correspondente (coluna segment_id de 01_road_segments.csv)
""")
